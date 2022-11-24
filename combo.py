
import pandas as pd
import codecs
import xlsxwriter
from pandas import ExcelWriter
from openpyxl import load_workbook
import os

# https://stackoverflow.com/questions/68629315/split-csv-file-into-multiple-data-frame-based-on-common-headers-in-rows-python

f = codecs.open("data.csv", "r",encoding="utf-8")
new_file = codecs.open("new_data.csv", "w",encoding="utf-8")
sheet_header = []
for lines in f.readlines():
    if not ",,,,," in lines:
        new_file.write(lines)
    elif ",,,,," in lines:
        sheet_header.append(lines.replace(",,,,,\r\n",""))

df = pd.read_excel("data.xlsx")
#filter all line which is header
split_lines = df[df['order']=='order']

dfs = []
last_idx = 0
for idx,row in split_lines.iterrows():
    #split line with the index
    _df = df[last_idx:idx]
    last_idx = idx+1
    dfs.append(_df)

_df = df[last_idx:df.shape[0]]
dfs.append(_df)

# Create Excel Workbook
LOG_File = os.path.join('TRACK_LOGS.xlsx')
workbook = xlsxwriter.Workbook(LOG_File)
worksheet = workbook.add_worksheet()
workbook.close()

# Load excel Workbook using openpyxl
book = load_workbook(LOG_File)
writer = ExcelWriter(LOG_File, engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

for df_indx,sub_df in enumerate(dfs):
    dfs[df_indx].to_excel(writer, sheet_name=sheet_header[df_indx], index=False)

if len(book.sheetnames) > 1:
    first_sheet = book['Sheet1']
    book.remove(first_sheet)

writer.save()